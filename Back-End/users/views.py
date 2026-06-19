from django.contrib.auth import get_user_model
from django.contrib.auth.password_validation import validate_password as dj_validate_password
from drf_spectacular.utils import extend_schema, extend_schema_view, inline_serializer
from rest_framework import serializers as s
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.viewsets import ModelViewSet

from audit_logs.services import AuditLogService
from brands.models import Brand
from common.permissions import ModulePermission
from common.responses import error_response, success_response
from common.utils import get_client_ip
from roles.models import Role

from .serializers import UserCreateSerializer, UserListSerializer, UserUpdateSerializer
from .services import UserService

User = get_user_model()


@extend_schema_view(
    list=extend_schema(summary='List all users', tags=['Users']),
    retrieve=extend_schema(summary='Get a user by ID', tags=['Users']),
    create=extend_schema(summary='Create a new user', tags=['Users']),
    update=extend_schema(summary='Update a user', tags=['Users']),
    partial_update=extend_schema(summary='Partially update a user', tags=['Users']),
    destroy=extend_schema(summary='Delete a user', tags=['Users']),
    activate=extend_schema(summary='Activate a user', tags=['Users'], request=None),
    deactivate=extend_schema(summary='Deactivate a user', tags=['Users'], request=None),
    reset_password=extend_schema(
        summary="Reset a user's password (admin)",
        tags=['Users'],
        request=inline_serializer(
            'ResetPasswordRequest',
            fields={'new_password': s.CharField(help_text='New password for the user')},
        ),
    ),
)
class UserViewSet(ModelViewSet):
    permission_classes  = [IsAuthenticated]
    queryset            = User.objects.select_related('role').prefetch_related('brands').all()
    search_fields       = ['username', 'role__name']   # 'mobile' removed — field doesn't exist
    ordering_fields     = ['created_at', 'is_active']
    # Note: role is filtered manually by role__name (not FK id) — see list()
    filterset_fields    = ['is_active']

    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        if self.action in ('update', 'partial_update'):
            return UserUpdateSerializer
        return UserListSerializer

    def get_permissions(self):
        action_map = {
            'list': 'view',
            'retrieve': 'view',
            'create': 'create',
            'update': 'edit',
            'partial_update': 'edit',
            'destroy': 'delete',
            'activate': 'activate',
            'deactivate': 'activate',
            'reset_password': 'edit',
            'bulk_import': 'create',
        }
        return [IsAuthenticated(), ModulePermission('users', action_map.get(self.action, 'view'))()]

    # ------------------------------------------------------------------
    # List
    # ------------------------------------------------------------------
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        # Accept both role id (?role=1) and role name (?role=rm).
        role_param = (request.query_params.get('role') or '').strip()
        if role_param:
            if role_param.isdigit():
                queryset = queryset.filter(role_id=int(role_param))
            else:
                queryset = queryset.filter(role__name__iexact=role_param.replace(' ', '_'))
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = UserListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        return success_response('Users fetched successfully', UserListSerializer(queryset, many=True).data)

    # ------------------------------------------------------------------
    # Retrieve
    # ------------------------------------------------------------------
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        return success_response('User fetched successfully', UserListSerializer(instance).data)

    # ------------------------------------------------------------------
    # Create
    # ------------------------------------------------------------------
    def create(self, request, *args, **kwargs):
        serializer = UserCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        AuditLogService.log(
            user=request.user,
            action='Created user',
            module='User',
            new_data=UserListSerializer(user).data,
            ip_address=get_client_ip(request),
        )
        return success_response('User created successfully', UserListSerializer(user).data, status_code=status.HTTP_201_CREATED)

    # ------------------------------------------------------------------
    # Update
    # ------------------------------------------------------------------
    def update(self, request, *args, **kwargs):
        partial  = kwargs.pop('partial', False)
        instance = self.get_object()
        old_data = UserListSerializer(instance).data
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        AuditLogService.log(
            user=request.user,
            action='Updated user',
            module='User',
            old_data=dict(old_data),
            new_data=UserListSerializer(user).data,
            ip_address=get_client_ip(request),
        )
        return success_response('User updated successfully', UserListSerializer(user).data)

    # ------------------------------------------------------------------
    # Destroy
    # ------------------------------------------------------------------
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if request.user.pk == instance.pk:
            return error_response('You cannot delete your own account.')
        AuditLogService.log(
            user=request.user,
            action='Deleted user',
            module='User',
            old_data={'id': instance.id, 'username': instance.username},
            ip_address=get_client_ip(request),
        )
        try:
            instance.delete()
        except Exception:
            # FK constraint — deactivate instead of hard-deleting
            instance.is_active = False
            instance.save(update_fields=['is_active', 'updated_at'])
            return success_response(
                'User has linked records and could not be permanently deleted. '
                'The account has been deactivated instead.',
                status_code=status.HTTP_200_OK,
            )
        return success_response('User deleted successfully', status_code=status.HTTP_204_NO_CONTENT)

    # ------------------------------------------------------------------
    # Custom actions
    # ------------------------------------------------------------------
    @action(detail=True, methods=['post'], url_path='activate')
    def activate(self, request, pk=None):
        user = self.get_object()
        if user.is_active:
            return error_response('User is already active')
        UserService.activate(user)
        AuditLogService.log(
            user=request.user,
            action='Activated user',
            module='User',
            new_data={'username': user.username},
            ip_address=get_client_ip(request),
        )
        return success_response('User activated successfully', UserListSerializer(user).data)

    @action(detail=True, methods=['post'], url_path='deactivate')
    def deactivate(self, request, pk=None):
        user = self.get_object()
        if request.user.pk == user.pk:
            return error_response('You cannot deactivate your own account.')
        if not user.is_active:
            return error_response('User is already inactive')
        UserService.deactivate(user)
        AuditLogService.log(
            user=request.user,
            action='Deactivated user',
            module='User',
            new_data={'username': user.username},
            ip_address=get_client_ip(request),
        )
        return success_response('User deactivated successfully', UserListSerializer(user).data)

    @action(detail=True, methods=['post'], url_path='reset-password')
    def reset_password(self, request, pk=None):
        from common.validators import validate_password_strength
        from rest_framework.exceptions import ValidationError as DRFValidationError
        user = self.get_object()
        new_password = request.data.get('new_password', '').strip()
        if not new_password:
            return error_response('new_password is required')
        try:
            validate_password_strength(new_password)
            dj_validate_password(new_password, user=user)
        except DRFValidationError as exc:
            return error_response(
                'Password validation failed',
                errors={'new_password': exc.detail if hasattr(exc, 'detail') else [str(exc)]},
            )
        except Exception as exc:
            return error_response(
                'Password validation failed',
                errors={'new_password': list(getattr(exc, 'messages', [str(exc)]))},
            )
        UserService.reset_password(user, new_password)
        AuditLogService.log(
            user=request.user,
            action='Reset user password',
            module='User',
            new_data={'username': user.username},
            ip_address=get_client_ip(request),
        )
        return success_response('Password reset successfully')

    @action(detail=False, methods=['post'], url_path='bulk-import')
    def bulk_import(self, request):
        """Bulk create users from a CSV or Excel file.

        Expected columns: username, role, brands (comma-separated), password (optional)
        Missing password defaults to 123456 and must_change_password is set to True.
        """
        import csv
        import io
        import openpyxl

        file = request.FILES.get('file')
        if not file:
            return error_response('File is required', status_code=status.HTTP_400_BAD_REQUEST)

        fname = file.name.lower()
        try:
            if fname.endswith('.csv'):
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                rows = [r for r in reader]
            elif fname.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(cell.value).strip() if cell.value is not None else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
            else:
                return error_response('Only CSV (.csv) and Excel (.xlsx) files are supported', status_code=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            return error_response(f'Failed to parse file: {exc}', status_code=status.HTTP_400_BAD_REQUEST)

        roles_cache  = {}
        for r in Role.objects.filter(is_active=True):
            # Match by name (lowercased) and also with spaces replaced by underscores
            roles_cache[r.name.lower()] = r
            roles_cache[r.name.lower().replace('_', ' ')] = r
            roles_cache[r.name.lower().replace(' ', '_')] = r

        brands_cache = {}
        for b in Brand.objects.all():
            # Match by exact name (case-insensitive)
            brands_cache[b.name.lower()] = b
            brands_cache[b.name.lower().replace(' ', '')] = b

        created, skipped, errors = [], [], []

        for i, row in enumerate(rows, start=2):
            username = str(row.get('username') or '').strip()
            if not username:
                errors.append({'row': i, 'error': 'Username is required'})
                continue

            if User.objects.filter(username=username).exists():
                skipped.append(username)
                continue

            role_key  = str(row.get('role') or '').strip().lower().replace(' ', '_')
            role_obj  = roles_cache.get(role_key)

            brand_names = [b.strip().lower() for b in str(row.get('brands') or '').split(',') if b.strip()]
            brand_objs  = []
            for bn in brand_names:
                brand = brands_cache.get(bn) or brands_cache.get(bn.replace(' ', ''))
                if brand:
                    brand_objs.append(brand)

            password = str(row.get('password') or '').strip() or '123456'

            if not role_obj and role_key:
                errors.append({'row': i, 'username': username, 'error': f"Role '{row.get('role', '')}' not found"})
                continue

            try:
                user = User(username=username, role=role_obj, must_change_password=True)
                user.set_password(password)
                user.save()
                if brand_objs:
                    user.brands.set(brand_objs)
                elif brand_names:
                    # Some brands weren't found — warn but still create user
                    unmatched = [bn for bn in brand_names if not (brands_cache.get(bn) or brands_cache.get(bn.replace(' ', '')))]
                    if unmatched:
                        errors.append({'row': i, 'username': username, 'error': f"Brand(s) not found: {', '.join(unmatched)} — user created without them"})
                created.append(username)
            except Exception as exc:
                errors.append({'row': i, 'username': username, 'error': str(exc)})

        AuditLogService.log(
            user=request.user,
            action=f'Bulk imported users: {len(created)} created, {len(skipped)} skipped, {len(errors)} errors',
            module='User',
            ip_address=get_client_ip(request),
        )
        return success_response('Bulk import completed', {
            'created': created,
            'skipped': skipped,
            'errors':  errors,
            'summary': f'{len(created)} created, {len(skipped)} skipped, {len(errors)} errors',
        })
